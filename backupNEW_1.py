import os
import sys
import telebot
from telebot import types
from openpyxl import load_workbook
import groups
from groupsButton import *


bot_token = "6497800204:AAFzMDXRKLX6DmYeaWIhYKSrTksTj1fQwaM"  # Replace YOUR_TOKEN_HERE with your actual bot token
bot = telebot.TeleBot(bot_token)



data = []
msg_txt = ''

def read_excel_file(file_path, sheet_name):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
     
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    return data

def back_to_start(message):
    bot.send_message(message.chat.id, "Вернёмся к началу...")
    start(message)

@bot.message_handler(commands=['start'])
def start(message):
    global data
    data = []
    buttons = ["понедельник", "вторник", "среда", "четверг", "пятница", "суббота"]
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(*map(types.KeyboardButton, buttons))
    bot.send_message(message.chat.id, text="какой сегодня день", reply_markup=markup)
    bot.register_next_step_handler(message, callback=select_course)
    

@bot.message_handler(func=lambda message: message.text in ["понедельник", "вторник", "среда", "четверг", "пятница", "суббота"])
def select_course(message):
    global msg_txt
    msg_txt = message.text
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("1 курс", "2 курс", "3 курс", "4 курс", "5 курс")
    bot.send_message(message.chat.id, text="Привет, Я бот для просмотра расписания.\n\nКакой у тебя курс?", reply_markup=markup)
    


@bot.message_handler(func=lambda message: message.text == "1 курс")
def onegroup(message):
    buttons = groups.buttons_course_1
    markup = types.InlineKeyboardMarkup()
    for index, button_text in enumerate(buttons, start=1):
        button = types.InlineKeyboardButton(button_text, callback_data=str(index))
        markup.add(button)

    bot.send_message(message.chat.id, text="Выбери свою группу 1 курса", reply_markup=markup)

@bot.message_handler(func=lambda message: message.text == "2 курс")
def onegroup(message):

    buttons = groups.buttons_course_2
    markup = types.InlineKeyboardMarkup()
    for index, button_text in enumerate(buttons, start=1):
        button = types.InlineKeyboardButton(button_text, callback_data=str(index + 22))
        markup.add(button)

    bot.send_message(message.chat.id, text="Выбери свою группу 2 курса", reply_markup=markup)

@bot.message_handler(func=lambda message: message.text == "3 курс")
def onegroup(message):

    buttons = groups.buttons_course_3
    markup = types.InlineKeyboardMarkup()
    for index, button_text in enumerate(buttons, start=1):
        button = types.InlineKeyboardButton(button_text, callback_data=str(index + 43))
        markup.add(button)

    bot.send_message(message.chat.id, text="Выбери свою группу 3 курса", reply_markup=markup)

@bot.message_handler(func=lambda message: message.text == "4 курс")
def onegroup(message):

    buttons = groups.buttons_course_4
    markup = types.InlineKeyboardMarkup()
    for index, button_text in enumerate(buttons, start=1):
        button = types.InlineKeyboardButton(button_text, callback_data=str(index + 64))
        markup.add(button)

    bot.send_message(message.chat.id, text="Выбери свою группу 4 курса", reply_markup=markup)

@bot.message_handler(func=lambda message: message.text == "5 курс")
def onegroup(message):

    buttons = groups.buttons_course_5
    markup = types.InlineKeyboardMarkup()
    for index, button_text in enumerate(buttons, start=1):
        button = types.InlineKeyboardButton(button_text, callback_data=str(index + 78))
        markup.add(button)

    bot.send_message(message.chat.id, text="Выбери свою группу 5 курса", reply_markup=markup)

@bot.callback_query_handler(func=lambda call: True)
def on_day_group_selected(call):
    if call.data == "back":
        back_to_start(call.message)
    else:
        global msg_txt
        sheet_name = msg_txt
        r_num = int(call.data) - 1 + 1
        data = read_excel_file('C:\\Users\\Observer\\Desktop\\uksap_bot\\UKSAP\\rasp.xlsx', sheet_name)
        current_data = data[r_num] 
        current_data_string = "\n\n".join(map(str, current_data))
        back_button = types.InlineKeyboardButton("Back", callback_data="back")
        markup = types.InlineKeyboardMarkup()
        markup.row(back_button)
        bot.send_message(call.message.chat.id, text=f"The data is:\n{current_data_string}", reply_markup=markup)

bot.polling()