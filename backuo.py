import os
import sys
import telebot
from telebot import types
from openpyxl import load_workbook


bot_token = "6497800204:AAFzMDXRKLX6DmYeaWIhYKSrTksTj1fQwaM"  # Replace YOUR_TOKEN_HERE with your actual bot token
bot = telebot.TeleBot(bot_token)

#days = ["понедельник", "вторник", "среда", "четверг", "пятница", "суббота"]


data = []
msg_txt = ''

def read_excel_file(file_path, sheet_name):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
     
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    return data

@bot.message_handler(commands=['start'])
def start(message):
    global data
    data = []
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    pn = types.KeyboardButton("понедельник")
    vt = types.KeyboardButton("вторник")
    sd = types.KeyboardButton("среда")
    ch = types.KeyboardButton("четверг")
    pt = types.KeyboardButton("пятница")
    sb = types.KeyboardButton("суббота")

    markup.add(pn, vt, sd, ch, pt, sb)
    bot.send_message(message.chat.id, text="какой сегодня день", reply_markup=markup)
    bot.register_next_step_handler(message, callback=select_course)
    

@bot.message_handler(func=lambda message: message.text in ["понедельник", "вторник", "среда", "четверг", "пятница", "суббота"])
def select_course(message):
    global msg_txt
    msg_txt = message.text
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("1 курс", "2 курс", "3 курс", "4 курс", "5 курс")
    bot.send_message(message.chat.id, text="Привет, Я бот для просмотра расписания.\n\nКакой у тебя курс?", reply_markup=markup)
    


@bot.message_handler(func=lambda message: message.text =="1 курс")
def onegroup(message):
    markup = types.InlineKeyboardMarkup()
    #НА БАЗЕ 9 КЛАССОВ
    A1239 = types.InlineKeyboardButton("А-123/9", callback_data ="1")
    A2239 = types.InlineKeyboardButton("А-223/9", callback_data="2")
    A3239 = types.InlineKeyboardButton("А-323/9", callback_data="3")
    D1239 = types.InlineKeyboardButton("Д-123/9", callback_data="4")
    D2239 = types.InlineKeyboardButton("Д-223/9", callback_data="5")
    D3239 = types.InlineKeyboardButton("Д-323/9", callback_data="6")
    C1239 = types.InlineKeyboardButton("С-123/9", callback_data="7")
    C2239 = types.InlineKeyboardButton("С-223/9", callback_data="8")
    ZM1239 = types.InlineKeyboardButton("ЗМ-123/9", callback_data="9")
    ZM2239 = types.InlineKeyboardButton("ЗМ-223/9", callback_data="10")
    IS1239 = types.InlineKeyboardButton("ИС-123/9", callback_data="11")
    IS2239 = types.InlineKeyboardButton("ИС-223/9", callback_data="12")
    M1239 = types.InlineKeyboardButton("М-123/9", callback_data="13")
    TG1239 = types.InlineKeyboardButton("ТГ-123/9", callback_data="14")
    TG2239 = types.InlineKeyboardButton("ТГ-223/9", callback_data="15")
    L1239 = types.InlineKeyboardButton("Л-123/9", callback_data="16")

    #НА БАЗЕ 11 КЛАССОВ
    A42311 = types.InlineKeyboardButton("А-423/11", callback_data="17")
    GD12311 = types.InlineKeyboardButton("ГД-123/11", callback_data="18")
    GD22311 = types.InlineKeyboardButton("ГД-223/11", callback_data="19")
    GD32311 = types.InlineKeyboardButton("ГД-323/11", callback_data="20")
    C32311 = types.InlineKeyboardButton("С-323/11", callback_data="21")
    ED12311 = types.InlineKeyboardButton("ЭД-123/11", callback_data="22")

    markup.add(A1239, A2239, A3239, D1239, D2239, D3239, C1239, C2239, ZM1239, ZM2239, IS1239, IS2239, M1239, TG1239, TG2239, L1239, ED12311)
    bot.send_message(message.chat.id, text="Выбери свою группу 1 курса", reply_markup=markup)


def get_data(call, r_num, sheet_name):
    data = read_excel_file('C:\\Users\\Observer\\Desktop\\TelegramBot\\rasp.xlsx', sheet_name)
    current_data = data[r_num] 
    current_data_string = "\n\n".join(map(str, current_data))    
    bot.send_message(call.message.chat.id, text=f"The data is:\n{current_data_string}")

@bot.callback_query_handler(func=lambda call: True)
def on_day_group_selected(call):
    global msg_txt
    sheet_name = msg_txt
    r_num = int(call.data) - 1
    get_data(call, r_num, sheet_name)

bot.polling()