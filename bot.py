import telebot
from telebot import types
from openpyxl import load_workbook
import groups
from find_info import choose_group, choose_day
from api_token import BOT_TOKEN, FILE_PATH


bot_token = BOT_TOKEN
bot = telebot.TeleBot(bot_token)

data = []
days = ["понедельник", "вторник", "среда", "четверг", "пятница", "суббота"]
courses = ["1 курс", "2 курс", "3 курс", "4 курс", "5 курс"]
first_course = ["2 курс", "3 курс", "4 курс", "5 курс"]
msg_txt = ''

def read_excel_file(file_path, sheet_name):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
     
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    return data

def back_to_start(message):
    bot.send_message(message.chat.id, "Вернёмся к началу...")
    start(message)

@bot.message_handler(commands=['start'])
def start(message):
    global data
    data = []
    buttons = days
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(*map(types.KeyboardButton, buttons))
    bot.send_message(message.chat.id, text="какой сегодня день", reply_markup=markup)
    bot.register_next_step_handler(message, callback=select_course)
    
@bot.message_handler(func=lambda message: message.text in days)
def select_course(message):
    global msg_txt
    if message.text == "понедельник":
        choose_day(message, types, bot, courses)
        msg_txt = message.text
    elif message.text == "вторник":
        choose_day(message, types, bot, courses)
        msg_txt = message.text
    elif message.text == "среда":
        choose_day(message, types, bot, courses)
        msg_txt = message.text
    elif message.text == "четверг":
        choose_day(message, types, bot, courses)
        msg_txt = message.text
    elif message.text == "пятница":
        choose_day(message, types, bot, courses)
        msg_txt = message.text    
    elif message.text == "суббота":
        choose_day(message, types, bot, first_course)
        msg_txt = message.text

@bot.message_handler(func=lambda message: message.text)
def onegroup(message):
    if message.text == "1 курс":
        choose_group(message, types, groups.buttons_course_1, bot, 0)
    elif message.text == "2 курс":
        choose_group(message, types, groups.buttons_course_2, bot, 22)
    elif message.text == "3 курс":
        choose_group(message, types, groups.buttons_course_3, bot, 43)
    elif message.text == "4 курс":
        choose_group(message, types, groups.buttons_course_4, bot, 64)
    elif message.text == "5 курс":
        choose_group(message, types, groups.buttons_course_5, bot, 78)

@bot.callback_query_handler(func=lambda call: True)
def on_day_group_selected(call):
    if call.data == "back":
        back_to_start(call.message)
    else:
        global msg_txt
        sheet_name = msg_txt
        r_num = int(call.data)
        data = read_excel_file(FILE_PATH, sheet_name)
        current_data = data[r_num] 
        current_data_string = "\n\n".join(map(str, current_data))
        back_button = types.InlineKeyboardButton("Назад", callback_data="back")
        markup = types.InlineKeyboardMarkup()
        markup.row(back_button)
        bot.send_message(call.message.chat.id, text=f"The data is:\n{current_data_string}", reply_markup=markup)

bot.polling()